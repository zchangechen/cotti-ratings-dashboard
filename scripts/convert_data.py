#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将 Excel/CSV 数据转换为仪表盘所需的 JSON 文件。"""

import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = SCRIPT_DIR / "data"
DOWNLOADS_DIR = SCRIPT_DIR / "downloads"

# 数据源路径（运行时会从仓库根目录读取）
# 格式：repo_root / source_path → (output_type, output_name)
# source_path 可以是相对路径（相对于脚本目录）
SOURCE_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else SCRIPT_DIR.parent / "_cotti_repo"


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for i, h in enumerate(headers):
            val = ws.cell(r, i + 1).value
            if val is None:
                row[h] = ""
            elif isinstance(val, float):
                # 保留合理精度
                row[h] = round(val, 6)
            else:
                row[h] = str(val).strip()
        rows.append(row)
    wb.close()
    return headers, rows


def read_csv(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = [row for row in reader]
    return headers, rows


def convert_stores(source_dir):
    """转换评分抓取结果 → stores.json"""
    xlsx_path = source_dir / "库迪咖啡_Google评分评论_抓取结果.xlsx"
    if not xlsx_path.exists():
        print(f"WARNING: 未找到 {xlsx_path}")
        return []

    _, rows = read_xlsx(xlsx_path)
    stores = []
    for row in rows:
        rating_raw = row.get("当前Google评分", "")
        reviews_raw = row.get("当前Google评论数", "")
        try:
            rating = float(rating_raw) if rating_raw else None
        except (ValueError, TypeError):
            rating = None
        try:
            reviews = int(reviews_raw) if reviews_raw else None
        except (ValueError, TypeError):
            reviews = None

        store = {
            "country": row.get("国家", ""),
            "store": row.get("门店", ""),
            "rating": rating,
            "reviews": reviews,
            "googleName": row.get("Google门店名称", ""),
            "googleLink": row.get("Google Maps 链接", ""),
            "status": row.get("查询状态", ""),
            "note": row.get("备注", ""),
        }
        stores.append(store)
    return stores


def convert_countries(stores):
    """根据 stores 数据生成国家汇总 → countries.json"""
    country_map = defaultdict(lambda: {
        "count": 0,
        "totalRating": 0,
        "ratingCount": 0,
        "totalReviews": 0,
        "bestRating": None,
        "bestStore": "",
        "worstRating": None,
        "worstStore": "",
    })

    for s in stores:
        c = s["country"]
        if not c:
            continue
        d = country_map[c]
        d["count"] += 1
        if s["rating"] is not None:
            d["totalRating"] += s["rating"]
            d["ratingCount"] += 1
            if d["bestRating"] is None or s["rating"] > d["bestRating"]:
                d["bestRating"] = s["rating"]
                d["bestStore"] = s["store"]
            if d["worstRating"] is None or s["rating"] < d["worstRating"]:
                d["worstRating"] = s["rating"]
                d["worstStore"] = s["store"]
        if s["reviews"] is not None:
            d["totalReviews"] += s["reviews"]

    countries = []
    for name, d in sorted(country_map.items()):
        countries.append({
            "name": name,
            "storeCount": d["count"],
            "avgRating": round(d["totalRating"] / d["ratingCount"], 2) if d["ratingCount"] > 0 else None,
            "ratedCount": d["ratingCount"],
            "totalReviews": d["totalReviews"],
            "bestRating": d["bestRating"],
            "bestStore": d["bestStore"],
            "worstRating": d["worstRating"],
            "worstStore": d["worstStore"],
        })
    return countries


def convert_problems(source_dir):
    """转换问题门店数据 → problems.json"""
    # 优先使用重分类版本
    for csv_name in ["问题门店验证结果_2026-06-02_重分类.csv", "问题门店验证结果_2026-06-02.csv"]:
        csv_path = source_dir / csv_name
        if csv_path.exists():
            _, rows = read_csv(csv_path)
            break
    else:
        print("WARNING: 未找到问题门店验证结果CSV")
        return []

    problems = []
    for row in rows:
        problems.append({
            "row": row.get("row", row.get("\ufeffrow", "")),
            "country": row.get("country", ""),
            "store": row.get("store", ""),
            "issueTypes": row.get("issue_types", ""),
            "currentStatus": row.get("current_status", ""),
            "currentNote": row.get("current_note", ""),
            "currentGoogleName": row.get("current_google_name", ""),
            "currentLink": row.get("current_link", ""),
            "scrapedStatus": row.get("scraped_status", ""),
            "scrapedRating": row.get("scraped_rating", ""),
            "scrapedReviews": row.get("scraped_reviews", ""),
            "suggestedStatus": row.get("suggested_status", ""),
            "suggestedNote": row.get("suggested_note", ""),
            "confidence": row.get("confidence", ""),
            "reason": row.get("reason", ""),
        })

    # 同时尝试读人工确认清单补充 priority
    manual_path = source_dir / "人工确认清单_2026-06-02.csv"
    if manual_path.exists():
        _, manual_rows = read_csv(manual_path)
        # 构建 (country, store) → priority 映射
        priority_map = {}
        for mr in manual_rows:
            key = (mr.get("country", ""), mr.get("store", ""))
            p = mr.get("priority", mr.get("\ufeffpriority", ""))
            if p:
                priority_map[key] = p

        for p in problems:
            key = (p["country"], p["store"])
            if key in priority_map:
                p["priority"] = priority_map[key]
            elif "高" in str(p.get("reason", "")):
                p["priority"] = "高"
            elif "低" in str(p.get("confidence", "")):
                p["priority"] = "低"
            else:
                p["priority"] = p.get("priority", "中")

    return problems


def convert_report(source_dir):
    """读取最新周报"""
    report_path = source_dir / "每周变化报告_2026-06-02.md"
    if not report_path.exists():
        # 找最新的
        reports = sorted(source_dir.glob("每周变化报告_*.md"), reverse=True)
        if reports:
            report_path = reports[0]
        else:
            return {"content": "暂无周报", "date": ""}

    with open(report_path, "r", encoding="utf-8") as f:
        content = f.read()

    # 提取日期
    date_str = report_path.stem.replace("每周变化报告_", "")
    return {"content": content, "date": date_str, "name": report_path.name}


def get_stats(stores, countries, problems):
    """计算全局统计"""
    rated_stores = [s for s in stores if s["rating"] is not None]
    all_reviews = sum(s["reviews"] for s in stores if s["reviews"] is not None)
    avg_rating = round(sum(s["rating"] for s in rated_stores) / len(rated_stores), 2) if rated_stores else 0

    # 评分分布
    distribution = {"0-3.0": 0, "3.0-3.5": 0, "3.5-4.0": 0, "4.0-4.5": 0, "4.5-5.0": 0}
    for s in rated_stores:
        r = s["rating"]
        if r < 3.0:
            distribution["0-3.0"] += 1
        elif r < 3.5:
            distribution["3.0-3.5"] += 1
        elif r < 4.0:
            distribution["3.5-4.0"] += 1
        elif r < 4.5:
            distribution["4.0-4.5"] += 1
        else:
            distribution["4.5-5.0"] += 1

    # 问题统计
    high_priority = sum(1 for p in problems if p.get("priority") == "高")
    mid_priority = sum(1 for p in problems if p.get("priority") == "中")
    low_priority = sum(1 for p in problems if p.get("priority") == "低")

    return {
        "totalStores": len(stores),
        "totalCountries": len(countries),
        "avgRating": avg_rating,
        "totalReviews": all_reviews,
        "ratedStoreCount": len(rated_stores),
        "ratingDistribution": distribution,
        "problemStats": {
            "total": len(problems),
            "high": high_priority,
            "mid": mid_priority,
            "low": low_priority,
        },
        "lastUpdated": "",
    }


def copy_downloads(source_dir):
    """复制可下载文件"""
    import shutil

    files_to_copy = [
        "门店信息库.xlsx",
        "库迪咖啡_Google评分评论_抓取结果.xlsx",
        "问题门店验证结果_2026-06-02.csv",
        "问题门店验证结果_2026-06-02_重分类.csv",
        "人工确认清单_2026-06-02.csv",
    ]
    copied = []
    for fname in files_to_copy:
        src = source_dir / fname
        dst = DOWNLOADS_DIR / fname
        if src.exists():
            shutil.copy2(src, dst)
            copied.append(fname)
    return copied


def main():
    source_dir = SOURCE_DIR
    print(f"数据源: {source_dir}")

    # 转换数据
    stores = convert_stores(source_dir)
    print(f"stores: {len(stores)} 条")

    countries = convert_countries(stores)
    print(f"countries: {len(countries)} 条")

    problems = convert_problems(source_dir)
    print(f"problems: {len(problems)} 条")

    report = convert_report(source_dir)
    print(f"report: {report.get('name', 'N/A')}")

    stats = get_stats(stores, countries, problems)

    # 写入 JSON
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

    for name, data in [
        ("stores.json", stores),
        ("countries.json", countries),
        ("problems.json", problems),
        ("report.json", report),
        ("stats.json", stats),
    ]:
        path = DATA_DIR / name
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"已写入: {path} ({len(json.dumps(data, ensure_ascii=False))} bytes)")

    # 复制可下载文件
    copied = copy_downloads(source_dir)
    print(f"已复制下载文件: {copied}")

    # 更新时间戳
    from datetime import datetime
    stats["lastUpdated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 更新 stores.json 中的 stats
    with open(DATA_DIR / "stores.json", "w", encoding="utf-8") as f:
        json.dump(stores, f, ensure_ascii=False, indent=2)

    print("\n转换完成！")


if __name__ == "__main__":
    main()
